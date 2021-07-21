from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

"""
-------
PYEXCEL
-------

AUTHOR: Tauanny Furlanetto.
OBJECTIVE: plot graphs from excel files using the openpyxl module.
FUNCTION: goes through an excel file, screens for a target (moving the desired  values to a new file), counts a specific value from the excel file,
adds the result and the labels to the excel documents, plots a pie chart showing the graph, the labels and the percent of each value.
OUTPUT: an excel file without a graph and one with it.

VERSION V. 0.1: Core functionalities developed, its still dependent of a specific file to work (it uses pre-defined information for parsing the file) it needs to be 
generalized and modularized. It also needs to be improved for higher memory and time efficiency.

"""

# ADDS THE ORIGINAL WORKBOOK
wb = load_workbook('pyExcel.xlsx')
ws = wb.active

#CREATES A NEW WORKBOOK
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Correct Values"


# CREATES A SERIES OF COUNTERS
c = 1 # Goes through the columns
i = 1 # Goes through the lines
j = 0 # Counts the amount of values that correspond to a certain target (in the 'dado' dict)
d = 1 # Goes through the data in the 'dado' dictionary

# CREATES A SERIES OF VARIABLES
coluna = 'F' # Defines the target column to be screened
valor = 'Sim' # Defines the value to be looked in the screening of coluna

deleteRows = [] # Defines the rows to be deleted (innactive)

valueCount = [] # Defines the amount of the values to be plotted in a graph
dado = {
	'1': 'Mulher',
	'2': 'Homem',
	'3': 'Não binário'
	} # Defines the values to be counted in the coluna2
coluna2 = 'C' # Defines the column to be counted


# MOVES THE DESIRED VALUES TO ANOTHER FILE
while c<(ws.max_column + 1):
	while i < (ws.max_row +1):
		if (ws[coluna+str(i)].value) == valor:
			ws2[str(get_column_letter(c))+str(i)] = ws[str(get_column_letter(c))+str(i)].value 
		elif c == 1:
			deleteRows.append(i)
		i = i + 1
	if i == (ws.max_row + 1): 
		i = 1
	c = c + 1


# COUNTS A SPECIFIC VALUE IN THE THE TABLE
while i<(ws2.max_row):
	if (ws2[coluna2+str(i)].value) == dado[str(d)]:
		print('==',d,'\n', i, '\n',j)
		j = j +1
	i = i +1
	if i == (ws2.max_row) and d<len(dado):
		valueCount.append(j) # IT HAS FOUND THE '3' VALUE, JUST CAN'T APPEND BECAUSE OF THE CONDITION
		j = 0
		i = 1
		d = d +1
	elif d == len(dado) and i==(ws2.max_row):

		valueCount.append(j)


# ADDS THE VALUE OF EACH COLUMN TO THE TABLE
i = 2 + ws2.max_row
for d in range(0, len(valueCount)):
	ws2[coluna2 +str(i)] = valueCount[d]
	i = i+1
	print(valueCount[d], i)


#ADDS THE COLUMN NAMES TO THE TABLE
for d in range(1, len(dado)+1):
	print(dado[str(d)])
	ws2.append({coluna2:dado[str(d)]})


#CREATES A FILE WITH THE VALUES COUNTED
wb2.save('correctValue.xlsx') 


# CREATES THE CHART
values = Reference (ws2, min_col = 3, min_row = (i-len(dado)),
					max_col = 3, max_row = int(i-1))
cats = Reference(ws2, min_col = 3, min_row=(i), max_row = int(i+2))
chart = PieChart()
chart.add_data(values)
chart.set_categories(cats)
chart.title = 'Genero de Designers'

# Customs the chart to show the data in the image
chart.title.delete = False
chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = 1
chart.dataLabels.showCatName = 0
chart.dataLabels.showVal = 0

#chart.x_axis.title = 'Gênero' - Inactive because of the type of chart
#chart.y_axis.title =  'Quantidade' - Inactive because of the type of chart

#ADDS THE CHART AND CREATES A NEW FILE WITH THE CHART IN IT
ws2.add_chart(chart, str(get_column_letter(5+ws2.max_column)+'1'))

wb2.save('grafico.xlsx')


